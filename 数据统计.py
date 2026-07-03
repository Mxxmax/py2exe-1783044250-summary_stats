"""
基础数据表.xlsx — 统计 + 未达账数据拆分
功能：
  1. 按检验师分组求和 & 占比（写入「求和」「占比」sheet）
  2. 筛选"未达账"行，按类别/船名/检验项目/港口/委托方/检验师 拆分为独立 .xlsx 文件
"""
import re, os, sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INPUT = sys.argv[1] if len(sys.argv) > 1 else "./基础数据.xlsx"
OUT_DIR = os.path.dirname(os.path.abspath(INPUT)) if os.path.dirname(INPUT) else "."
SUM_SHEET = "求和"
PCT_SHEET = "占比"

# ── 配置 ──
TARGET_NAMES = [  # 求和/占比用的数值列
    "确认收（CNY）", "公车", "私车", "交通费", "过路费",
    "住宿费", "住船费", "补贴", "入港费", "防疫费", "登轮费",
    "带教", "假期补助", "其他", "实际住宿费", "检验天数",
]
SPLIT_COLS = ["类别", "船名", "检验项目", "港口", "委托方", "检验师"]  # 拆分列
ARRIVAL_COL = "到账日期"  # 用于筛选"未达账"

# ── 工具 ──
def normalize(name):
    return re.sub(r"\s+", "", str(name))

def build_header_map(ws):
    """读取表头，返回 normalized_name -> 0-based index"""
    raw = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    hmap = {}
    for i, h in enumerate(raw):
        key = normalize(h)
        if key not in hmap:
            hmap[key] = i
    return hmap, raw

def lookup(hmap, name):
    """按名称查列索引，支持括号半角/全角容错"""
    key = normalize(name)
    idx = hmap.get(key)
    if idx is None:
        alt = key.replace("（", "(").replace("）", ")")
        idx = hmap.get(alt)
    return idx

def safe_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def is_weida(input_date):
    """判断是否等于"未达账"等表示未到账的字符串"""
    if input_date is None or isinstance(input_date, (int, float)):
        return False
    s = str(input_date).strip()
    return s in ("未达账", "未达账", "未到账", "未到账")

# ── 样式 ──
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
total_fill = PatternFill("solid", fgColor="D9E2F3")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def write_data_rows(ws, headers_raw, rows, col_indices):
    """将 rows 写入 ws（带表头和样式），col_indices 是需输出的列索引列表"""
    # 表头
    for ci, idx in enumerate(col_indices, 1):
        cell = ws.cell(row=1, column=ci, value=headers_raw[idx])
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    # 数据
    for ri, row in enumerate(rows, 2):
        for ci, idx in enumerate(col_indices, 1):
            cell = ws.cell(row=ri, column=ci, value=row[idx])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if isinstance(row[idx], (int, float)) else "left")
    # 列宽
    for ci in range(1, len(col_indices) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(ws.cell(row=r, column=ci).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


# ══════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════

wb = openpyxl.load_workbook(INPUT, data_only=True)
ws = wb.active
hmap, headers_raw = build_header_map(ws)
ncols = len(headers_raw)

# ── 读取全量行数据（按行号索引） ──
all_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
all_rows = [r for r in all_rows if any(c is not None for c in r)]  # 跳过全空行

# ══════════════════════════════════════════════════
# 功能1: 求和 + 占比
# ══════════════════════════════════════════════════
inspector_idx = lookup(hmap, "检验师")
if inspector_idx is None:
    print(f"❌ 找不到列「检验师」")
    exit(1)

target_cols = {}
for name in TARGET_NAMES:
    idx = lookup(hmap, name)
    if idx is None:
        print(f"⚠️ 找不到列「{name}」，跳过")
        continue
    target_cols[name] = idx

if target_cols:
    # 提取数据
    stat_data = []
    for row in all_rows:
        ins = str(row[inspector_idx]).strip() if row[inspector_idx] else ""
        if not ins:
            continue
        vals = {name: safe_float(row[idx]) for name, idx in target_cols.items()}
        stat_data.append((ins, vals))

    # 分组求和
    groups = {}
    for ins, vals in stat_data:
        groups.setdefault(ins, {col: 0.0 for col in target_cols})
        for col in target_cols:
            groups[ins][col] += vals[col]

    sorted_ins = sorted(groups.keys())
    totals = {col: sum(groups[ins][col] for ins in sorted_ins) for col in target_cols}

    def write_summary(sheet, is_pct):
        headers = ["检验师分组"] + list(target_cols.keys())
        n = len(headers)
        for ci, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        for ri, ins in enumerate(sorted_ins, 2):
            sheet.cell(row=ri, column=1, value=ins).border = thin_border
            sheet.cell(row=ri, column=1).alignment = Alignment(horizontal="left")
            for ci, col in enumerate(target_cols, 2):
                if is_pct:
                    total = totals[col]
                    val = groups[ins][col]
                    pct = (val / total * 100) if total != 0 else 0.0
                    cell = sheet.cell(row=ri, column=ci, value=round(pct, 2))
                    cell.number_format = '0.00"%"'
                else:
                    cell = sheet.cell(row=ri, column=ci, value=round(groups[ins][col], 4))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right")
        tr = len(sorted_ins) + 2
        cell = sheet.cell(row=tr, column=1, value="合计")
        cell.font = Font(bold=True, size=11)
        cell.fill = total_fill
        cell.border = thin_border
        for ci, col in enumerate(target_cols, 2):
            cell = sheet.cell(row=tr, column=ci, value=100.0 if is_pct else round(totals[col], 4))
            if is_pct:
                cell.number_format = '0.00"%"'
            cell.font = Font(bold=True, size=11)
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
        for ci in range(1, n + 1):
            col_letter = get_column_letter(ci)
            ml = max(len(str(sheet.cell(row=r, column=ci).value or "")) for r in range(1, tr + 1))
            sheet.column_dimensions[col_letter].width = min(ml + 4, 22)

    for name in [SUM_SHEET, PCT_SHEET]:
        if name in wb.sheetnames:
            del wb[name]
    write_summary(wb.create_sheet(SUM_SHEET), is_pct=False)
    write_summary(wb.create_sheet(PCT_SHEET), is_pct=True)

    OUTPUT = INPUT.replace(".xlsx", "_统计结果.xlsx")
    wb.save(OUTPUT)
    print(f"✅ 统计结果 → {OUTPUT}")
    print(f"   共 {len(sorted_ins)} 个检验师分组")
else:
    print("⚠️ 未找到数值列，跳过统计功能")


# ══════════════════════════════════════════════════
# 功能2: 筛选未达账 + 按列拆分
# ══════════════════════════════════════════════════
arrival_idx = lookup(hmap, ARRIVAL_COL)
if arrival_idx is None:
    print(f"⚠️ 找不到列「{ARRIVAL_COL}」，跳过拆分功能")
    sys.exit(0)

# 过滤出未达账行
weida_rows = [row for row in all_rows if is_weida(row[arrival_idx])]

if not weida_rows:
    print("⚠️ 没有未达账数据，跳过拆分")
    sys.exit(0)

# 所有输出列：整行所有列（保持原始表头）
all_col_indices = list(range(ncols))

for split_name in SPLIT_COLS:
    split_idx = lookup(hmap, split_name)
    if split_idx is None:
        print(f"⚠️ 找不到列「{split_name}」，跳过")
        continue

    # 按该列值分组
    groups_split = {}
    for row in weida_rows:
        val = str(row[split_idx]).strip() if row[split_idx] else "(空)"
        groups_split.setdefault(val, []).append(row)

    # 建目录
    dir_path = os.path.join(OUT_DIR, split_name)
    os.makedirs(dir_path, exist_ok=True)

    for key, rows in groups_split.items():
        # 文件名：去掉非法字符
        safe_key = re.sub(r'[\\/:*?"<>|]', "_", key)
        fpath = os.path.join(dir_path, f"{safe_key}.xlsx")
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        write_data_rows(out_ws, headers_raw, rows, all_col_indices)
        out_wb.save(fpath)

    print(f"📁 {split_name}/  → {len(groups_split)} 个文件")

print(f"✅ 未达账行共 {len(weida_rows)} 条，拆分完成")
